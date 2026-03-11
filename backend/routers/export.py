import csv
import io
from datetime import date, datetime
from typing import Any

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from auth import get_current_user
from database import get_db
from models import Case, User

router = APIRouter(prefix="/api/export", tags=["export"], redirect_slashes=False)

# Column definitions: (header_label, case_attribute)
EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("Case ID", "id"),
    ("Patient ID", "patient_id"),
    ("Date", "case_date"),
    ("Hospital", "hospital"),
    ("Clinical Role", "clinical_role"),
    ("Category", "category"),
    ("Type", "type"),
    ("Procedure", "procedure"),
    ("Detail", "detail"),
    ("Observations", "obs"),
    ("Outcome", "outcome"),
    ("Pregnant", "pregnant"),
    ("Complications", "complications"),
    ("Previous CS", "prev_cs"),
    ("Sterilisation", "sterilisation"),
    ("Pregnancy Check Date", "pregnancy_check_date"),
    ("Oocyte Data", "oocyte_data"),
    ("ET Data", "et_data"),
    ("Extra Data", "extra_data"),
    ("Created At", "created_at"),
]


def _format_value(value: Any) -> str:
    """Convert a model value to a plain string for export."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    if isinstance(value, dict):
        import json
        return json.dumps(value, default=str)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


async def _fetch_user_cases(
    user: User, db: AsyncSession
) -> list[Case]:
    result = await db.execute(
        select(Case)
        .where(Case.user_id == user.id)
        .order_by(Case.case_date.desc().nullslast(), Case.created_at.desc())
    )
    return list(result.scalars().all())


# ── CSV export ────────────────────────────────────────────────────────────────
@router.get("/csv")
async def export_csv(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    cases = await _fetch_user_cases(current_user, db)

    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)

    # Header row
    writer.writerow([col[0] for col in EXPORT_COLUMNS])

    # Data rows
    for case in cases:
        writer.writerow(
            [_format_value(getattr(case, attr)) for _, attr in EXPORT_COLUMNS]
        )

    output.seek(0)

    filename = f"prolog_cases_{current_user.username}_{date.today().isoformat()}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Excel export ──────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


@router.get("/excel")
async def export_excel(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    cases = await _fetch_user_cases(current_user, db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cases"  # type: ignore[assignment]

    headers = [col[0] for col in EXPORT_COLUMNS]
    ws.append(headers)  # type: ignore[union-attr]

    # Style header row
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)  # type: ignore[union-attr]
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for row_idx, case in enumerate(cases, start=2):
        row_data = [
            _format_value(getattr(case, attr)) for _, attr in EXPORT_COLUMNS
        ]
        ws.append(row_data)  # type: ignore[union-attr]

        # Alternate row shading
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _ALT_FILL  # type: ignore[union-attr]

    # Auto-size columns (cap at 60 chars wide)
    for col_idx, (header, _) in enumerate(EXPORT_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(header),
            *(
                len(_format_value(getattr(case, attr)))
                for case in cases
                for h, attr in [(EXPORT_COLUMNS[col_idx - 1])]
            ),
            0,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)  # type: ignore[union-attr]

    # Freeze header row
    ws.freeze_panes = "A2"  # type: ignore[union-attr]

    # Write to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = (
        f"prolog_cases_{current_user.username}_{date.today().isoformat()}.xlsx"
    )
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
